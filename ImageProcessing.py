from PIL import Image
from PyQt5.QtWidgets import QDesktopWidget, QWidget
from openpyxl import Workbook, load_workbook
import os

from Preprocessing import Preprocessing
from WarningWindow import WarningWindow
import math



class ImageProcessing:

    def __init__(self):
        self.preProc = Preprocessing()

    def findFirstChanOnImg(chanY: int, chanN: int, pixOfMark: int, yBeg: int, yDest: int, compYPix: int):
        ymin = 155 + (compYPix - 1) * 3
        ymin = max(ymin, yBeg)
        if pixOfMark < yBeg:
            while pixOfMark < ymin:
                pixOfMark += chanY
                chanN += 1
        else:
            while pixOfMark - chanY > ymin:
                pixOfMark -= chanY
                chanN -= 1
        if pixOfMark<yBeg or pixOfMark>yDest:
            return -1, -1
        return chanN, pixOfMark


    def measurement(self, xBeg: int, yBeg: int, yDest: int, app: QWidget):
        """Funkcja służąca pomierzeniu uszkodzeń i zapisaniu ich do pliku"""

        mmPerPix = self.preProc.readNumber(43, 96, 120, 120, app, 'x')
        compYPix = self.preProc.readNumber(355, 96, 397, 120, app, 'compY Pixels')
        x_scale_val, x_scale_pos = self.preProc.findBeltX()
        markedChan, pixLine, chanY = self.preProc.findBeltChan()

        self.preProc.binarization('markedArea.png', 'binarizated.png', 200)

        #ustalenie x dla poczatku screena wzgledem osi
        CONVERTER = 0.67    #zmienna przeliczająca mmPerPix na liczbę milimetrów odpowiadających 1 pikselowi
        x_scale_val *= 1000
        x_scale_val += round(((xBeg - x_scale_pos)*mmPerPix)*CONVERTER)
        chanN, pixLine = ImageProcessing.findFirstChanOnImg(chanY, markedChan, pixLine, yBeg, yDest, compYPix)
        if pixLine < 0:
            WarningWindow('Blad odczytu osi Y! Pomiar nie wykonany!')
            return
        pixLine = pixLine - yBeg

        # otwieranie pliku
        if os.path.isfile('pomiaryUszkodzen.xlsx') and os.access('pomiaryUszkodzen.xlsx', os.W_OK):
            wb = load_workbook('pomiaryUszkodzen.xlsx')
            ws = wb.active
            ws.title = "dane"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "dane"
            headingRow = []
            headingRow.append("Kanal")
            headingRow.append("Poczatek [m]")
            headingRow.append("Kolor")
            headingRow.append("Uszkodzenie [mm]")
            headingRow.append("Poczatek [m]")
            headingRow.append("Kolor")
            headingRow.append("Uszkodzenie [mm]")
            headingRow.append("Poczatek [m]")
            headingRow.append("Kolor")
            headingRow.append("Uszkodzenie [mm]")
            ws.append(headingRow)
        binarizated = Image.open('binarizated.png')
        width, height = binarizated.size
        yellow = True           # opisuje, czy aktualnie mierzymy obszar niebieski, czy zolty
        countingOn = False      # opisuje, czy poprzedni piksel byl czarny (zbinaryzowany szary), czy nie, a wiec czy jest wlaczone zliczanie
        greenCounting = 0       # opisuje, czy poprzedni piksel byl zielony
        damageLen = 0
        redCounting = 0
        showedWarning = False   # czy wyswietlono ostrzezenie o ucieciu paska
        showedWarning2 = False  # czy wyswietleno ostrzezenie o zaznaczeniu obszaru z nieprzestawionym wskaznikiem
        sheetRow = []
        emptyChan = False
        ymax = QDesktopWidget().screenGeometry().height()*2 - 90
        ymax = min(yDest, ymax) # ostatni wiersz pikseli na ktorym nalezy wykonac pomiary
        while pixLine + yBeg < ymax:
            # Jezeli poprzedni wiersz konczyl sie w pasku, a nie w "szarej strefie"
            if countingOn:
                if yellow == True:
                    # Wpisanie pozycji poczatku uszkodzenia
                    sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                    # Wpisanie koloru uszkodzenia
                    sheetRow.append('zolty')
                    sheetRow.append('z ' + str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                else:
                    # Wpisanie pozycji poczatku uszkodzenia
                    sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                    # Wpisanie koloru uszkodzenia
                    sheetRow.append('niebieski')
                    sheetRow.append('n ' + str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                damageLen = 0
                countingOn = False
                greenCounting = 0
                redCounting = 0
                if not showedWarning:
                    WarningWindow('Mozliwe uciecie paska! Mozliwie zle zapisane dane!')
                    showedWarning = True
            if sheetRow != []:
                ws.append(sheetRow)
                emptyChan = False
            else:
                # Jezeli poprzedni kanal nie byl pusty to zrobi pusty wiersz w Excelu
                if emptyChan == False:
                    ws.append(sheetRow)
                emptyChan = True
            sheetRow.clear()
            firstDamageOnChan = True
            for x in range(width):
                p = binarizated.getpixel((x, pixLine))
                if p[1] == 255:
                    # Jesli kolor zolty
                    if p[0] == 255 and p[2] == 0:
                        if x == 0 and not showedWarning:
                            WarningWindow('Mozliwe uciecie paska! Mozliwie zle zapisane dane!')
                            showedWarning = True
                        if firstDamageOnChan:
                            firstDamageOnChan = False
                            # Numer kanalu
                            sheetRow.append(str(chanN))
                        # Przed zoltym jest czerwony
                        if redCounting > 0:
                            if yellow:
                                damageLen += redCounting + 1
                            else:
                                yellow = True
                            redCounting = 0
                        # Przed zoltym jest tez piksel zolty
                        elif countingOn == True and yellow == True and greenCounting == 0:
                            damageLen += 1
                        # Przed zoltym jest niebieski
                        elif countingOn == True and yellow == False and greenCounting == 0:
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('niebieski')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = 1
                            yellow = True
                        # Przed zoltym jest szary
                        elif countingOn == False:
                            countingOn = True
                            damageLen = 1
                            yellow = True
                        # Przed zoltym jest zielony
                        else:
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('niebieski')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = greenCounting + 1
                            greenCounting = 0
                            yellow = True
                    # Jesli kolor niebieski
                    elif p[0] == 0 and p[2] == 255:
                        if x == 0 and not showedWarning:
                            WarningWindow('Mozliwe uciecie paska! Mozliwie zle zapisane dane!')
                            showedWarning = True
                        if firstDamageOnChan:
                            firstDamageOnChan = False
                            sheetRow.append(str(chanN))
                        if redCounting > 0:
                            if yellow == False:
                                damageLen += redCounting + 1
                            else:
                                yellow = False
                            redCounting = 0
                        # Przed niebieskim jest tez piksel niebieski
                        elif countingOn == True and yellow == False and greenCounting == 0:
                            damageLen += 1
                        # Przed niebieskim jest zolty
                        elif countingOn == True and yellow == True and greenCounting == 0:
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('zolty')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = 1
                            yellow = False
                        # Przed niebieskim jest szary
                        elif countingOn == False:
                            countingOn = True
                            damageLen = 1
                            yellow = False
                        # Przed niebieskim jest zielony
                        else:
                            yellow = False
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('zolty')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = greenCounting + 1
                            greenCounting = 0
                    # kolor zielony
                    elif countingOn:
                        greenCounting += 1


                # kolor szary (lub jakis blad koloru czerwonego/niebieskiego)
                else:
                    # kolor czerwony
                    if p[0] == 255:
                        if not showedWarning2:
                            WarningWindow(
                                'Usun czerwony wskaznik z zaznaczonego pola! Dane nie zostaly zapisane!')
                            showedWarning2 = True
                            break
                        if countingOn:
                            redCounting += 1
                    else:
                        redCounting = 0
                    # zakonczenie paska...
                    if countingOn:
                        countingOn = False
                        # ...zoltego
                        if yellow == True:
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('zolty')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = 0
                            greenCounting = 0
                        # ...niebieskiego
                        else:
                            # Wpisanie pozycji poczatku uszkodzenia
                            sheetRow.append(str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * CONVERTER))/1000))
                            # Wpisanie koloru uszkodzenia
                            sheetRow.append('niebieski')
                            # Wpisanie dlugosci uszkodzenia
                            sheetRow.append(str(round(float((damageLen + greenCounting) * mmPerPix) * CONVERTER)))
                            damageLen = 0
                            greenCounting = 0
            pixLine += chanY
            chanN += 1
        try:
            if not showedWarning2:
                wb.save('pomiaryUszkodzen.xlsx')
        except Exception:
            WarningWindow('Prosze zamknac Excela przed rozpoczeciem pomiarow! Pomiar niezapisany!')
        return

    @staticmethod
    def binarizationMIN():
        img = Image.open('markedArea.png')
        width, height = img.size
        black_white = Image.new(mode="L", size=(width, height))

        for y in range(height):
            for x in range(width):
                p = img.getpixel((x, y))
                min_val = min(p[0], p[1], p[2])
                if min_val < 150:
                    min_val = 0
                else:
                    min_val = 255
                black_white.putpixel((x, y), min_val)
        black_white.save('chanels.png')

    #Funkcja wyliczająca rząd pierwszego kanału na obrazku.
    #Jeśli pierwszy rząd będzie poza obrazkiem, to zwróci -1
    @staticmethod
    def firstChan(yBeg: int, yDest: int, compYPix: int, chanY: int):
        chanN = math.ceil((154 - 3*compYPix - yBeg)/(-2*(chanY-1)))
        if chanN < yDest:
            chanN -= yBeg
            return chanN
        else:
            return -1

    # Funkcja wyliczająca rząd ostatniego kanału na obrazku.
    @staticmethod
    def lastChan(yBeg: int, yDest: int, compYPix: int, chanY: int):
        chanN = math.floor((154 - 3 * compYPix - yDest) / (-2 * (chanY - 1)))
        chanN -= yBeg
        return chanN